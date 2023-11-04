from board import *
from individual import *
from population import *
from openpyxl import *

def exitCB():
    global running, exiting
    running = False
    exiting = True


def manageExcel(population, mutation_rate, generation):
    filename = '8_queens.xlsx'
    wb = load_workbook(filename)

    ws = wb.create_sheet(f'Population{population}')

    wb.save(filename)

    ws.cell(row=1, column = 1, value = "Mutation rate")
    ws.cell(row=2, column = 1, value = mutation_rate)

    ws.cell(row=1, column = 2, value = "Generations needed")
    ws.cell(row=2, column = 2, value = generation)

    wb.save(filename)
    

def startAlgo():
    global running, exiting
    running = True

    pop = Population(int(populationInput.get()))
    pop.mutation_rate = mutationRate.get() / 100.0
    sliderText.set("Mutation rate: %.2f" % pop.mutation_rate)
    generation = 0
    while running and not exiting:
        genText.set("Generation: %d" % generation)
        pop.evaluate()
        pop.select_threshold()
        pop.crossover()
        pop.mutate()
        pop.individuals = pop.offspring
        i = pop.getBestIndividual()
        fitText.set("Fitness: %.2f" % i.fitness)
        wi = pop.getWorstIndividual()
        minFitText.set("Worst fitness: %.2f" % wi.fitness)
        board.setPos(i.dna)
        board.redraw(root)
        root.update_idletasks()
        root.update()
        if i.fitness == 1.0:
            running = False
            manageExcel(pop.size, pop.mutation_rate, generation)
        generation += 1
        

running = False
exiting = False

root = Tk()
root.title("PyGA")
root.minsize(524, 820)
root.maxsize(524, 820)

exitButton = Button(root, text="Exit", command = exitCB, font=("Arial", 25))
exitButton.place(x = 160, y = 750)

startButton = Button(root, text="Start", command = startAlgo, font=("Arial", 25))
startButton.place(x = 270, y = 750)

genText = StringVar()
generationLabel = Label( root, textvariable=genText, font=("Arial", 25))
generationLabel.place(x = 50, y = 532)

fitText = StringVar()
fitnessLabel = Label( root, textvariable=fitText, font=("Arial", 25))
fitnessLabel.place(x = 330, y = 592)

minFitText = StringVar()
minFitnessLabel = Label( root, textvariable=minFitText, font=("Arial", 25))
minFitnessLabel.place(x = 265, y = 532)

mutationRate = DoubleVar()
w2 = Scale( root, variable=mutationRate,from_=0, to=100,tickinterval=10, length=424, orient=HORIZONTAL)
w2.place(x = 50, y = 680)

sliderText = StringVar()
sliderLabel = Label( root, textvariable=sliderText, font=("Arial", 25))
sliderLabel.place(x = 50, y = 592)

populationLabel = Label( root, text="Population", font=("Arial", 25))
populationLabel.place(x = 50, y = 642)

populationInput = Entry(root, width = 30)
populationInput.place(x = 190, y = 647, height=32)

board = Board()

while not exiting:
    root.update_idletasks()
    root.update()

root.destroy